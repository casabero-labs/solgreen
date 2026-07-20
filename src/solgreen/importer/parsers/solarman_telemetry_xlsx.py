from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path

from openpyxl import load_workbook

from solgreen.contracts import SourceType
from solgreen.importer.exceptions import CorruptFileError
from solgreen.importer.parsers.base import BaseParser, ParsedFile


class SolarmanTelemetryXlsxParser(BaseParser):
    source_type = SourceType.SOLARMAN_INVERTER_TELEMETRY

    def parse(self, path: Path) -> ParsedFile:
        workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
        try:
            sheet = workbook.active
            if sheet is None:
                raise CorruptFileError(path, "no active sheet")
            rows_iter = sheet.iter_rows(values_only=True)
            header = next(rows_iter, None)
            if header is None:
                raise CorruptFileError(path, "empty sheet")
            columns = tuple(str(c).strip() if c is not None else "" for c in header)
            rows: list[dict[str, object]] = []
            for row in rows_iter:
                if row is None or all(c is None for c in row):
                    continue
                record: dict[str, object] = {}
                for col, value in zip(columns, row, strict=True):
                    record[col] = value
                rows.append(record)
            return ParsedFile(
                source_type=self.source_type,
                columns=columns,
                rows=rows,
            )
        finally:
            workbook.close()

    def iter_rows(self, path: Path) -> Iterator[dict[str, object]]:
        yield from self.parse(path).rows
