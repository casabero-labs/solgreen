from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path

from openpyxl import load_workbook

from solgreen.contracts import SourceType
from solgreen.importer.exceptions import CorruptFileError, HeaderMismatchError
from solgreen.importer.parsers.base import PLANT_FLOW_COLUMNS, BaseParser, ParsedFile


class SolarmanFlowXlsxParser(BaseParser):
    source_type = SourceType.SOLARMAN_PLANT_FLOW

    def parse(self, path: Path) -> ParsedFile:
        workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
        try:
            sheet = workbook.active
            if sheet is None:
                raise CorruptFileError(path, "no active sheet")
            rows_iter = sheet.iter_rows(values_only=True)
            header = next(rows_iter, None)
            if header is None:
                raise CorruptFileError(path, "empty sheet")
            columns = tuple(str(c).strip() if c is not None else "" for c in header)
            self._validate_headers(path, columns)
            rows: list[dict[str, object]] = []
            for row in rows_iter:
                if row is None or all(c is None for c in row):
                    continue
                rows.append(self._row_to_dict(columns, row))
            return ParsedFile(
                source_type=self.source_type,
                columns=columns,
                rows=rows,
            )
        finally:
            workbook.close()

    def iter_rows(self, path: Path) -> Iterator[dict[str, object]]:
        yield from self.parse(path).rows

    @staticmethod
    def _validate_headers(path: Path, observed: tuple[str, ...]) -> None:
        missing = tuple(c for c in PLANT_FLOW_COLUMNS if c not in observed)
        if missing:
            raise HeaderMismatchError(path=path, missing=missing, unexpected=())

    @staticmethod
    def _row_to_dict(columns: tuple[str, ...], row: tuple[object, ...]) -> dict[str, object]:
        record: dict[str, object] = {}
        for col, value in zip(columns, row, strict=True):
            record[col] = value
        return record
