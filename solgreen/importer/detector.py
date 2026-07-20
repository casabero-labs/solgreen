from pathlib import Path

from openpyxl import load_workbook
from polars import read_csv

from solgreen.contracts import SourceType

PLANT_FLOW_REQUIRED: tuple[str, ...] = (
    "Nombre de planta",
    "Hora actualizada",
    "Zona horaria",
    "Potencia de producción(W)",
    "SoC(%)",
)
PLANT_FLOW_COLUMN_COUNT = 12

INVERTER_TELEMETRY_REQUIRED: tuple[str, ...] = (
    "número de serie",
    "Hora actualizada",
    "PV Open Circuit Voltage(V)",
    "Voltaje CC PV1(V)",
    "Grid Code",
    "BUS voltage(V)",
)


def _normalize_columns(columns: list[str] | tuple[str, ...]) -> tuple[str, ...]:
    return tuple(c.strip() for c in columns)


def _read_csv_header(path: Path) -> tuple[str, ...]:
    try:
        frame = read_csv(path, n_rows=0)
    except Exception:
        return ()
    return _normalize_columns(frame.columns)


def _read_xlsx_header(path: Path) -> tuple[str, ...]:
    try:
        workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception:
        return ()
    try:
        sheet = workbook.active
        if sheet is None:
            return ()
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return ()
        return _normalize_columns([str(c) if c is not None else "" for c in header_row])
    finally:
        workbook.close()


def detect_format(path: Path) -> SourceType:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        columns = _read_csv_header(path)
    elif suffix in {".xlsx", ".xlsm"}:
        columns = _read_xlsx_header(path)
    else:
        return SourceType.UNKNOWN

    if not columns:
        return SourceType.UNKNOWN

    column_set = set(columns)
    if all(col in column_set for col in INVERTER_TELEMETRY_REQUIRED):
        return SourceType.SOLARMAN_INVERTER_TELEMETRY
    if (
        all(col in column_set for col in PLANT_FLOW_REQUIRED)
        and len(columns) == PLANT_FLOW_COLUMN_COUNT
    ):
        return SourceType.SOLARMAN_PLANT_FLOW
    return SourceType.UNKNOWN
